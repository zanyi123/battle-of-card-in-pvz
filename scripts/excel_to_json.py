"""excel_to_json.py — 从Excel源数据重新生成正确的 cards.json。

映射策略：
- Excel名称 → 去除后缀(fa/fu/s/t)得到纯中文名
- image_file: 基于阵营前缀和语义匹配已有图片文件
- effect_id: 基于Excel效果描述匹配游戏内效果代码
- description: 直接使用Excel效果描述
"""
import json
from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path(r"C:\Users\Lenovo\Desktop\pvzcard-infor.xlsx")
OUTPUT_PATH = Path(r"e:\cursor储存\pvz-plant-card-game\config\cards.json")
ASSETS_DIR = Path(r"e:\cursor储存\pvz-plant-card-game\assets\cards")

# ═══════════════════════════════════════════════════════════════
# Excel名称 → image_file 的精确映射（人工确认）
# 格式: "Excel中的完整名称（含后缀）": "image_file"
# ═══════════════════════════════════════════════════════════════
NAME_TO_IMAGE: dict[str, str] = {
    # ── 辅(fu) 15张 ──────────────────────────────────────
    "向日葵fu":      "fu_01_xiangri.png",
    "爆裂葡萄fu":     "fu_08_baolie.png",
    "阳光蓓蕾fu":     "fu_10_yangguang.png",
    "巴豆fu":        "fu_15_badou.png",
    "仙桃fu":        "fu_16_xiantao.png",
    "棉小雪fu":      "fu_17_mainxiao.png",
    "竹笋fu":        "fu_22_zhusun.png",
    "龙舌兰fu":      "fu_26_longshe.png",
    "阳光菇fu":      "fu_27_yangguangg.png",
    "香水蘑菇fu":     "fu_38_xiangshui.png",
    "粉丝心叶兰fu":   "fu_39_fensi.png",
    "魅惑菇fu":      "fu_40_meihuo.png",
    "能量花fu":      "fu_41_nengliang.png",
    "向日葵歌手fu":   "fu_42_geshou.png",
    "三叶草fu":      "fu_43_sanye.png",
    # ── 射(sh) 13张 ──────────────────────────────────────
    "豌豆射手s":      "sh_02_wandou.png",
    "梅小美s":        "sh_05_meixiao.png",
    "玉米投手s":      "sh_06_yumi.png",
    "寒冰射手s":      "sh_09_hanbing.png",
    "幽暮投手s":      "sh_11_youmu.png",
    "梛子加农炮s":    "sh_21_yezi.png",
    "仙人掌s":        "sh_28_xianren.png",
    "红针花s":        "sh_29_hongzhen.png",
    "莲小蓬s":        "sh_30_lainxiao.png",
    "小喷菇s":        "sh_31_xiaopen.png",
    "飞镖洋蓟s":      "sh_32_feibiao.png",
    "西瓜投手s":      "sh_33_xigua.png",
    "双重射手s":      "sh_54_shuangfa.png",
    # ── 坦(tk) 13张 ──────────────────────────────────────
    "坚果墙t":        "tk_03_jianguo.png",
    "原始猕猴桃t":    "tk_07_yuanshihou.png",
    "原始坚果墙t":    "tk_12_yuanshijian.png",
    "窝瓜t":          "tk_13_wogua.png",
    "炙热山葵t":      "tk_14_chire.png",
    "白萝卜t":        "tk_18_bailuo.png",
    "旋转菠萝t":      "tk_36_xuanzhuan.png",
    "榴莲t":          "tk_37_liulain.png",
    "熊果臼炮t":      "tk_44_xiongguo.png",
    "胆小荆棘t":      "tk_45_danxiao.png",
    "暗樱草t":        "tk_46_anying.png",
    "南瓜头t":        "tk_47_nangua.png",
    "全息坚果t":      "tk_48_quanxi.png",
    # ── 法(fa) 13张 ──────────────────────────────────────
    "火爆辣椒fa":      "fa_04_huobao.png",
    "冰龙草fa":        "fa_19_binglong.png",
    "魔术菇fa":        "fa_20_moshu.png",
    "火龙果fa":        "fa_23_huolongc.png",
    "火龙草fa":        "fa_23_huolongg.png",
    "闪电芦苇fa":      "fa_25_shandian.png",
    "毁灭菇fa":        "fa_34_huimie.png",
    "毒影菇fa":        "fa_35_duying.png",
    "南瓜巫师fa":      "fa_49_nangwu.png",
    "强酸柠檬fa":      "fa_50_qiangsuan.png",
    "热辣海藻fa":      "fa_51_rela.png",
    "岩浆番石榴fa":    "fa_52_yanjiang.png",
    "水晶兰fa":        "fa_53_shuijing.png",
}

# ═══════════════════════════════════════════════════════════════
# Excel效果描述 → effect_id 的映射
# ═══════════════════════════════════════════════════════════════
EFFECT_MAP: dict[str, str] = {
    "抵挡一回合攻击":                         "SHIELD_TURN",
    "扺挡一回合攻击":                         "SHIELD_TURN",
    "将对手手牌中的法师阵营卡随机抽出作废（有播报）": "DISCARD_FA",
    "护盾+6":                                "SHIELD_6",
    "将血量恢复至8点":                        "HEAL_8",
    "使对方卡牌攻击失效（仅攻击）":            "ATK_DISABLE",
    "沉默：该回合对手无法出牌":                "SILENCE",
    "破甲：无视护盾":                         "ARMOR_PIERCE",
    "对克制它阵营卡牌的攻击值清0":            "COUNTER_ATK_ZERO",
    "他对克制的阵营伤害×3":                   "COUNTER_DMG_X3",
    "若无法被克制，伤害×2":                   "NO_COUNTER_DMG_X2",
    "将对手手牌中的射手阵营卡随机抽出添加到自己手牌": "STEAL_SH",
    "从对面手牌抽一张加入自己手牌":            "STEAL_CARD",
    "将同出的卡牌消耗费用值转为回复血量":       "COST_TO_HEAL",
    "将对方攻击反弹对方":                     "REFLECT_ATK",
    "增加同出卡牌伤害2点，恢复2点生命":        "BOOST_ATK_HEAL",
    "对对手卡牌每张减伤2点":                  "REDUCE_DMG_2",
    "恢复：自己消耗费用等于生命恢复值":         "COST_TO_HEAL_SELF",
    "将对方手牌攻击值转化为恢复血量":           "ATK_TO_HEAL",
}

# 后缀 → 阵营代码
SUFFIX_TO_FACTION: dict[str, str] = {
    "fa": "法",
    "fu": "辅",
    "s": "射",
    "t": "坦",
}


def _parse_effect(effect_raw: str) -> str:
    """将Excel效果描述映射为 effect_id。"""
    if not effect_raw or not effect_raw.strip():
        return ""
    cleaned = effect_raw.strip()
    # Excel中效果文本常含换行符，去除换行和空格后再匹配
    normalized = cleaned.replace("\n", "").replace(" ", "")
    return EFFECT_MAP.get(normalized, "")


def _strip_suffix(name: str) -> tuple[str, str]:
    """从名称中提取阵营后缀并返回(纯名称, 阵营代码)。"""
    for suffix, faction in SUFFIX_TO_FACTION.items():
        if name.endswith(suffix):
            return name[: -len(suffix)].strip(), faction
    return name.strip(), ""


def main() -> None:
    # ── 读取Excel ──────────────────────────────────────────
    wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb.active
    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # 跳过表头
            continue
        if not any(row):
            continue
        rows.append(list(row))
    wb.close()

    # ── 列映射 ────────────────────────────────────────────
    # 0=植物名字, 2=攻击值, 4=费用消耗, 8=限制符(1为有), 11=特殊技能
    COL_NAME = 0
    COL_ATK = 2
    COL_COST = 4
    COL_LIMIT = 8
    COL_EFFECT = 11

    # ── 去重并构建卡牌 ────────────────────────────────────
    seen_names: set[str] = set()
    cards: list[dict] = []
    for row in rows:
        excel_name = str(row[COL_NAME]).strip() if row[COL_NAME] else ""
        if not excel_name or excel_name in seen_names:
            continue
        seen_names.add(excel_name)

        clean_name, faction = _strip_suffix(excel_name)
        cost = int(row[COL_COST]) if row[COL_COST] is not None else 0
        atk = int(row[COL_ATK]) if row[COL_ATK] is not None else 0
        limit_flag = str(row[COL_LIMIT]).strip() == "1" if row[COL_LIMIT] else False
        effect_raw = str(row[COL_EFFECT]).strip() if row[COL_EFFECT] else ""
        effect_id = _parse_effect(effect_raw)
        image_file = NAME_TO_IMAGE.get(excel_name, "")

        card = {
            "id": len(cards) + 1,
            "name": clean_name,
            "cost": cost,
            "atk": atk,
            "faction": faction,
            "type": "主",
            "limit_flag": limit_flag,
            "effect_id": effect_id,
            "description": effect_raw if effect_raw else "",
            "image_file": image_file,
        }
        cards.append(card)

    # ── 验证图片完整性 ────────────────────────────────────
    existing_imgs = {f.name for f in ASSETS_DIR.glob("*.png")}
    missing_imgs = []
    duplicate_imgs = set()
    used_imgs: set[str] = set()
    for c in cards:
        img = c["image_file"]
        if not img:
            missing_imgs.append(c["name"])
        elif img not in existing_imgs:
            missing_imgs.append(f"{c['name']} → {img}")
        elif img in used_imgs:
            duplicate_imgs.add(img)
        else:
            used_imgs.add(img)

    if missing_imgs:
        print(f"[WARN] Missing images: {missing_imgs}")
    if duplicate_imgs:
        print(f"[WARN] Duplicate images: {duplicate_imgs}")

    # ── 构建输出JSON ──────────────────────────────────────
    output = {
        "_comment": "由 excel_to_json.py 从 Excel 源数据自动生成，请勿手动编辑",
        "meta": {
            "total_cards": len(cards),
            "source": "pvzcard-infor.xlsx",
            "factions": {"法": "fa", "射": "sh", "坦": "tk", "辅": "fu"},
        },
        "field_docs": {
            "id": "卡牌唯一ID（1起始递增）",
            "name": "卡牌名称（纯中文，不含阵营后缀）",
            "cost": "法力消耗",
            "atk": "攻击值",
            "faction": "阵营（法/射/坦/辅）",
            "type": "卡牌类型（主）",
            "limit_flag": "限制符（true=限1张）",
            "effect_id": "技能ID（空=无技能）",
            "description": "技能描述（Excel原文）",
            "image_file": "图片文件名（assets/cards/下）",
        },
        "cards": cards,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"OK: generated {len(cards)} cards -> {OUTPUT_PATH}")
    fa_n = sum(1 for c in cards if c["faction"] == "\u6cd5")
    sh_n = sum(1 for c in cards if c["faction"] == "\u5c04")
    tk_n = sum(1 for c in cards if c["faction"] == "\u5766")
    fu_n = sum(1 for c in cards if c["faction"] == "\u8f85")
    print(f"     fa:{fa_n} sh:{sh_n} tk:{tk_n} fu:{fu_n}")

    # 打印有技能的卡牌
    skill_cards = [c for c in cards if c["effect_id"]]
    if skill_cards:
        print(f"\n=== 有技能的卡牌 ({len(skill_cards)}张) ===")
        for c in skill_cards:
            print(f"  {c['name']:10s} | {c['effect_id']:20s} | {c['description'][:40]}")


if __name__ == "__main__":
    main()
